from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from util.constant import PROPERTY_DASHBOARD_COLUMNS, PROPERTY_DETAIL_COLUMNS


class ExcelGenerator:
    def convert_lists_to_strings(self, df):
        """
        Convert lists in DataFrame cells to strings.

        :param df: DataFrame, DataFrame with potential lists in cells
        :return: DataFrame, DataFrame with lists converted to strings
        """
        for col in df.columns:
            df[col] = df[col].apply(lambda x: ', '.join(
                x) if isinstance(x, list) else x)
        return df

    def convert_feasible_to_yes_no(self, df):
        """
        Convert boolean values in the 'feasible' column to 'Yes' or 'No'.

        :param df: DataFrame, DataFrame with a 'feasible' column
        :return: DataFrame, DataFrame with 'feasible' values converted to 'Yes' or 'No'
        """
        if 'feasible' in df.columns:
            df['feasible'] = df['feasible'].apply(
                lambda x: 'Yes' if x else 'No')
        return df

    def format_percentage_columns(self, sheet, percentage_columns, col_name_to_letter):
        """
        Format specified columns as percentage in the given sheet.

        :param sheet: openpyxl worksheet, the worksheet to format
        :param percentage_columns: list of str, columns to format as percentage
        :param col_name_to_letter: dict, mapping of column names to column letters
        """
        percentage_style = NamedStyle(
            name="percentage_style", number_format="0.00%")
        for col in percentage_columns:
            col_letter = col_name_to_letter.get(col)
            if col_letter:
                for cell in sheet[col_letter]:
                    if cell.row > 1:  # Skip header row
                        cell.style = percentage_style

    def generate_dashboard(self, property_data):
        property_data = self.convert_lists_to_strings(property_data)
        property_data = self.convert_feasible_to_yes_no(property_data)

        # Ensure the 'feasible' column is after 'highest_margin'
        columns_order = [
            "address",
            "advertised_price",
            "size",
            "bed",
            "bath",
            "car",
            "state",
            "suburb",
            "postcode",
            "date_listed",
            "days_on_market",
            "a_subdivision_gross_profit_on_cost",
            "b_subdivision_reno_gross_profit_on_cost",
            "c_demolish_subdivision_gross_profit_on_cost",
            "d_demolish_duplex_gross_profit_on_cost",
            "e_subdivision_reno_duplex_gross_profit_on_cost",
            "f_demolish_townhouse_gross_profit_on_cost",
            "highest_margin",
            "feasible"  # Move 'feasible' after 'highest_margin'
        ]

        # Reorder columns in the DataFrame
        property_data = property_data[columns_order]

        # Rename columns
        property_data = property_data.rename(
            columns=PROPERTY_DASHBOARD_COLUMNS)

        output = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dashboard"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment = Alignment(horizontal="center", vertical="center")

        col_name_to_letter = {}
        for col_idx, col in enumerate(property_data.columns, 1):
            cell = sheet.cell(row=1, column=col_idx, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment
            col_name_to_letter[col] = cell.column_letter

        for row_idx, row in property_data.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=row_idx + 2, column=col_idx, value=value)
                cell.border = border

        percentage_columns = [
            PROPERTY_DASHBOARD_COLUMNS["a_subdivision_gross_profit_on_cost"],
            PROPERTY_DASHBOARD_COLUMNS["b_subdivision_reno_gross_profit_on_cost"],
            PROPERTY_DASHBOARD_COLUMNS["c_demolish_subdivision_gross_profit_on_cost"],
            PROPERTY_DASHBOARD_COLUMNS["d_demolish_duplex_gross_profit_on_cost"],
            PROPERTY_DASHBOARD_COLUMNS["e_subdivision_reno_duplex_gross_profit_on_cost"],
            PROPERTY_DASHBOARD_COLUMNS["f_demolish_townhouse_gross_profit_on_cost"],
            PROPERTY_DASHBOARD_COLUMNS["highest_margin"]
        ]

        self.format_percentage_columns(
            sheet, percentage_columns, col_name_to_letter)

        workbook.save(output)
        output.seek(0)
        return output

    def generate_property_detail(self, property_data, property_id):
        property_data = self.convert_lists_to_strings(property_data)

        # Map the 'Description' column to human-readable names
        property_data['Description'] = property_data['Description'].map(
            PROPERTY_DETAIL_COLUMNS)

        output = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"property_{property_id[:20]}"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment = Alignment(horizontal="center", vertical="center")

        sheet.cell(row=1, column=1, value="Description").font = header_font
        sheet.cell(row=1, column=2, value="Data").font = header_font
        sheet.cell(row=1, column=1).fill = header_fill
        sheet.cell(row=1, column=2).fill = header_fill
        sheet.cell(row=1, column=1).border = border
        sheet.cell(row=1, column=2).border = border
        sheet.cell(row=1, column=1).alignment = alignment
        sheet.cell(row=1, column=2).alignment = alignment

        # Identify percentage fields
        percentage_fields = [
            PROPERTY_DETAIL_COLUMNS["stamp_duty"],
            PROPERTY_DETAIL_COLUMNS["conveyancing_fees"],
            PROPERTY_DETAIL_COLUMNS["bank_legal_fees_on_purchase"],
            PROPERTY_DETAIL_COLUMNS["interest_on_residential_loan"],
            PROPERTY_DETAIL_COLUMNS["interest_on_construction_loan"],
            PROPERTY_DETAIL_COLUMNS["agents_commission"],
            PROPERTY_DETAIL_COLUMNS["conveyancing_settlement_costs"],
            PROPERTY_DETAIL_COLUMNS["gst_liability_on_sales"],
            PROPERTY_DETAIL_COLUMNS["professional_fees_based_on_construction_cost"],
            PROPERTY_DETAIL_COLUMNS["contingencies_based_on_build_cost"],
            PROPERTY_DETAIL_COLUMNS["bank_fees"],
            PROPERTY_DETAIL_COLUMNS["brokers_fees"],
            PROPERTY_DETAIL_COLUMNS["other_lending_costs"],
            PROPERTY_DETAIL_COLUMNS["investor_input"],
            PROPERTY_DETAIL_COLUMNS["renovation_uplift"],
            PROPERTY_DETAIL_COLUMNS["renovation_costs"],
            PROPERTY_DETAIL_COLUMNS["useable_land"],
            PROPERTY_DETAIL_COLUMNS["construction_deposit"]
        ]

        # Identify currency fields
        currency_fields = [
            PROPERTY_DETAIL_COLUMNS["advertised_price"],
            PROPERTY_DETAIL_COLUMNS["find_price"],
            PROPERTY_DETAIL_COLUMNS["lower_price"],
            PROPERTY_DETAIL_COLUMNS["mid_price"],
            PROPERTY_DETAIL_COLUMNS["upper_price"],
            PROPERTY_DETAIL_COLUMNS["net_rental_income_per_week"],
            PROPERTY_DETAIL_COLUMNS["civil_costs_per_block"],
            PROPERTY_DETAIL_COLUMNS["demolition_costs"],
            PROPERTY_DETAIL_COLUMNS["build_cost_m2"],
            PROPERTY_DETAIL_COLUMNS["council_contributions_per_house"],
            PROPERTY_DETAIL_COLUMNS["rates_utilities_land_tax"],
            PROPERTY_DETAIL_COLUMNS["insurance_on_existing_buildings"],
        ]

        percentage_style = NamedStyle(
            name="percentage_style", number_format="0.00%")
        currency_style = NamedStyle(
            name="currency_style", number_format='"$"#,##0')

        for row_idx, row in property_data.iterrows():
            desc_cell = sheet.cell(
                row=row_idx + 2, column=1, value=row['Description'])
            data_cell = sheet.cell(
                row=row_idx + 2, column=2, value=row['Data'])
            desc_cell.border = border
            data_cell.border = border

            if row['Description'] in percentage_fields:
                data_cell.style = percentage_style
            if row['Description'] in currency_fields:
                data_cell.style = currency_style

        workbook.save(output)
        output.seek(0)
        return output
